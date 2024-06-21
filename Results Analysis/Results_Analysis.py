#**************************************************************************** Evaluation Metrics Calculation for ProSmart HPA  ****************************************************************************

import math
import numpy as np
import statistics
import openpyxl
from openpyxl import Workbook, load_workbook


#**************************************************************************** Open Microservices Data Stored in Knowledge Base  ****************************************************************************

def open_files(filename, run):

    if run == 1:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 1/{filename}.xlsx')                                  # PREDICTION WINDOW SIZE can be 5, 10, 15, 20, 25, and 30 seconds
    elif run == 2:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 2/{filename}.xlsx')
    elif run == 3:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 3/{filename}.xlsx')
    elif run == 4:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 4/{filename}.xlsx')
    elif run == 5:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 5/{filename}.xlsx')
    elif run == 6:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 6/{filename}.xlsx')
    elif run == 7:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 7/{filename}.xlsx')
    elif run == 8:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 8/{filename}.xlsx')
    elif run == 9:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 9/{filename}.xlsx')
    elif run == 10:
        workbook = load_workbook(f'./PATH TO KNOWLEDGE BASE/PREDICTION WINDOW SIZE/Run 10/{filename}.xlsx')
    
    return workbook


#**************************************************************************** Main Code for stored data analysis  ****************************************************************************

                                   
replicas = 2                                       # Considered Experimental scenario (i.e., CPU Threshold 20% with 2 Max. Replicas)
CPU_Threshold = 20     

filename = "frontend"                              #Starting data analysis from the frontend microservice
CPU_Request = 100                                  #CPU Request value for frontend microservice
                                   
run = 1                                            #Starting from the first run analysis and continues till the 10th run
micro = 0                                          # micro represents microservice number: Starting from 1st till the 11th microservice                                    


# ************************************************************ Evaluation Metrics Array for storing each microservice data separately *****************************************************************
Supply_CPU =[]
CPU_OverUtilization = []
OverProvision_CPU = []
UnderProvision_CPU = []

# ************************************************************ Evaluation Metrics Array for storing each run data separately *****************************************************************
Ten_Supply_CPU =[]
Ten_CPU_OverUtilization = []
Ten_OverProvision_CPU = []
Ten_UnderProvision_CPU = []



# ******************************************* Starting analyzing results for each run separately, and inside each run analysis, analyzing evaluation metrics for each microservice separately  *****************************************************************

while (run <= 10):

    micro = micro + 1

    workbook = open_files (filename, run)
      
    worksheet = workbook.active

    #**************************************************************************** Getting Test Timestamps and corresponding CPU utilization values ********************************************************************

    test_time = []                                       #Getting load test time stamps from the stored data
    for cell in worksheet['A'][1:]:
        if cell.value is not None:
            test_time.append(cell.value)


    cpu_utilization = []                                 #Getting CPU Utilization (percentage) values
    for cell in worksheet['B'][1:]:
        if cell.value is not None:
            cpu_utilization.append(cell.value)
    Avg_CPU_Utilization = statistics.mean(cpu_utilization)



    #**************************************************************************** CPU Overutilization Calculation ********************************************************************


    
    overutilized_cpu = []                                 #Getting CPU Overutilization values where cpu_utilization > CPU_Threshold
    for i in range(len(cpu_utilization)):
        if cpu_utilization[i] > CPU_Threshold:
            overutilized_cpu.append(cpu_utilization[i] - CPU_Threshold)
        else:
            overutilized_cpu.append(0)


    Avg_overutilized_CPU = statistics.mean(overutilized_cpu)




    #**************************************************************************** Supply CPU Calculation ********************************************************************

    
    current_replicas = []
    for cell in worksheet['C'][1:]:
        if cell.value is not None:
            current_replicas.append(cell.value)



    Avg_used_allocated_cpu_list = [value * CPU_Request for value in current_replicas]

    Avg_used_allocated_cpu = statistics.mean(Avg_used_allocated_cpu_list)



    #**************************************************************************** Resource Capacity Calculation ******************************************************

    max_replicas = []
    for cell in worksheet['E'][1:]:
        if cell.value is not None:
            max_replicas.append(cell.value)



    Avg_max_available_cpu_list = [value * CPU_Request for value in max_replicas]

    Avg_max_available_cpu = statistics.mean (Avg_max_available_cpu_list)


    #**************************************************************************** Resource Demand Calculation ***********************************************************


    desired_replicas = []

    for cell in worksheet['D'][1:]:
        if cell.value is not None:
            desired_replicas.append(cell.value)


    Avg_desired_cpu_list = [value * CPU_Request for value in desired_replicas]

    Avg_desired_cpu = statistics.mean (Avg_desired_cpu_list)


    #**************************************************************************** Overprovisioned CPU Calculation **************************************************************

    Residual_replicas = []


    for i in range(len(max_replicas)):
        if max_replicas[i] > desired_replicas[i]:
            Residual_replicas.append(max_replicas[i] - desired_replicas[i])
        else:
            Residual_replicas.append(0)

    Avg_residual_cpu_list = [value * CPU_Request for value in Residual_replicas]

    Avg_residual_cpu = statistics.mean (Avg_residual_cpu_list)




    #**************************************************************************** Underprovisioned CPU Calculation ******************************************************

    required_replicas = []


    for i in range(len(max_replicas)):
        if max_replicas[i] < desired_replicas[i]:
            required_replicas.append(desired_replicas[i] - max_replicas[i])
        else:
            required_replicas.append(0)

    Avg_required_cpu_list = [value * CPU_Request for value in required_replicas]

    Avg_required_cpu = statistics.mean (Avg_required_cpu_list)
    

    #******************************************************** Storing Analyzed Results for each microservice in a Separate File ***********************************************************************

    workbook = load_workbook(f'./Knowledge Base/{filename}.xlsx')                              #filename is the microservice name

    worksheet = workbook.active
    r= 5                                       # row number of CSV file for storing results
    c= 0                                       # column number of CSV file for storing results

    if run == 1:
        c = 9
    elif run == 2:
        c = 11
    elif run == 3:
        c = 13
    elif run == 4:
        c = 15
    elif run == 5:
        c = 17
    elif run == 6:
        c = 19
    elif run == 7:
        c = 21
    elif run == 8:
        c = 23
    elif run == 9:
        c = 25
    elif run == 10:
        c = 27
    
    worksheet.cell(row=r, column=c, value=Avg_used_allocated_cpu)
    r = r + 1
  
    worksheet.cell(row=r, column=c, value=Avg_overutilized_CPU)
    r = r + 1
    
    worksheet.cell(row=r, column=c, value=Avg_residual_cpu)
    r = r + 1
    
    worksheet.cell(row=r, column=c, value=Avg_required_cpu)
    r = r + 1

    workbook.save(f'./Knowledge Base/{filename}.xlsx')


    #**************************************************** Storing Evaluation metrics for all microservices (one by one) ***************************************************************
    
    Supply_CPU.append(Avg_used_allocated_cpu)
    CPU_OverUtilization.append(Avg_overutilized_CPU)
    OverProvision_CPU.append(Avg_residual_cpu)
    UnderProvision_CPU.append(Avg_required_cpu)


    # ********************************************************************* Switching to Next Microservice ***********************************************************************
    
    if filename == "frontend":
        filename = "adservice"
        CPU_Request = 200
    
    elif filename == "adservice":
        filename = "cartservice"
        CPU_Request = 200

    elif filename == "cartservice":
        filename = "paymentservice"
        CPU_Request = 100

    elif filename == "paymentservice":
        filename = "currencyservice"
        CPU_Request = 100

    elif filename == "currencyservice":
        filename = "emailservice"
        CPU_Request = 100

    elif filename == "emailservice":
        filename = "checkoutservice"
        CPU_Request = 100

    elif filename == "checkoutservice":
        filename = "productcatalogservice"
        CPU_Request = 100
    
    elif filename == "productcatalogservice":
        filename = "recommendationservice"
        CPU_Request = 100

    elif filename == "recommendationservice":
        filename = "redis-cart"
        CPU_Request = 70

    elif filename == "redis-cart":
        filename = "shippingservice"
        CPU_Request = 100

    elif filename == "shippingservice":
        filename = "frontend"
        CPU_Request = 100
        
    

    #*************************************************** Collection of analyzed Evaluation Metrics for all microservices for a single run for the Application_Level Analysis later on *****************************************************
    
    if (micro == 11):   

        Final_Supply_CPU = sum(Supply_CPU)
        Final_CPU_OverUtilization = sum(CPU_OverUtilization)
        Final_OverProvision_CPU = sum(OverProvision_CPU)
        Final_UnderProvision_CPU = sum(UnderProvision_CPU)
    

        # Storing results for all 10 runs

        Ten_Supply_CPU.append(Final_Supply_CPU)
        Ten_CPU_OverUtilization.append(Final_CPU_OverUtilization)
        Ten_OverProvision_CPU.append(Final_OverProvision_CPU)
        Ten_UnderProvision_CPU.append(Final_UnderProvision_CPU)


        # initializing parameters for getting and analyzing data of next run

        run = run + 1
        micro = 0

        Supply_CPU = []
        CPU_OverUtilization = []
        OverProvision_CPU = []
        UnderProvision_CPU = []



#**************************************************************************************** Application Level Analysis for all 10 runs ***************************************************************************************  

def insert_array_values(roww, data_values, start_column):
    for i, value in enumerate(data_values):
        worksheet.cell(row=roww, column=start_column, value=value)
        start_column = start_column + 1
        


Final_Supply_CPU = statistics.mean(Ten_Supply_CPU)
Final_CPU_OverUtilization = statistics.mean(Ten_CPU_OverUtilization)
Final_OverProvision_CPU = statistics.mean(Ten_OverProvision_CPU)
Final_UnderProvision_CPU = statistics.mean(Ten_UnderProvision_CPU)



workbook = load_workbook(f'./Knowledge Base/Application Level Analysis.xlsx')

worksheet = workbook.active


insert_array_values(6, Ten_Supply_CPU, 2)
insert_array_values(7, Ten_CPU_OverUtilization, 2)
insert_array_values(8, Ten_OverProvision_CPU, 2)
insert_array_values(9, Ten_UnderProvision_CPU, 2)


workbook.save(f'./Knowledge Base/Application Level Analysis.xlsx')

