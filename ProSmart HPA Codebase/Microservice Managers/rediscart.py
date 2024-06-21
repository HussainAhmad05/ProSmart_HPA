
#***************************************************** This file outlines the algorithm for Microservice Manager responsible for the redis microservice **********************************************

import sys
import fnmatch
import glob, os
import subprocess
import math
import numpy as np
import time
import statistics
from openpyxl import Workbook, load_workbook
import pandas as pd
import pickle
from Application_Resource_Manager import *


#************************************************************************************ Monitor Component **********************************************************************************************

def Monitor ():

    microservice_name = "redis-cart"

    # Monitoring Current Replicas 
                                                   
    Available_Replicas = 1
    Operational_Replicas = 0

    while Available_Replicas != Operational_Replicas:               #To ensure available replicas are in operational state
    
        Available_Replicas = "kubectl get deployment redis-cart -o=jsonpath='{.status.availableReplicas}'"
        Available_Replicas = subprocess.check_output(Available_Replicas.split()).decode('utf-8')
        if Available_Replicas.strip("'").isdigit():
            Available_Replicas = int(Available_Replicas.strip("'"))
        else:
            Available_Replicas = 0
        
        try:
            Replicas_CPU_usage = 'kubectl top pods -l app=redis-cart'
            Replicas_CPU_usage = subprocess.check_output(Replicas_CPU_usage.split()).decode('utf-8')
            Operational_Replicas = max(0, len(Replicas_CPU_usage.splitlines()) - 1)              
        except subprocess.CalledProcessError:
            Operational_Replicas = 0
            Available_Replicas = 1

    current_replicas = Available_Replicas

    # Calculating total Averaged current CPU Utilization across current replica pods

    cpu_add = []
    for line in Replicas_CPU_usage.splitlines()[1:]:
        columns = line.split()
        cpu_usage = columns[1]
        cpu_usage = cpu_usage[:-1]
        cpu_add = cpu_add + [int(cpu_usage)]

    current_cpu = math.ceil(statistics.mean(cpu_add))

    # Monitoring CPU Request Value for a Replica Deployment

    cpu_request = "kubectl get deployment redis-cart -o=jsonpath='{.spec.template.spec.containers[0].resources.requests.cpu}'"
    cpu_request = subprocess.check_output(cpu_request.split()).decode('utf-8')
    cpu_request = cpu_request[:-2]
    cpu_request = int(cpu_request.strip("'"))

    # Setting SLA Metrics: CPU Threshold, Minimum Replicas, and Maximum Replicas

    target_cpu = 20                                        # Considered Experimental Scenario 2R-20% 
    max_replica = 2
    min_replica = 1

    return microservice_name, current_replicas, current_cpu, target_cpu, cpu_request, max_replica, min_replica



#***************************************************************************** Analyze Component *****************************************************************************************


def Analyse (Test_Time, current_replicas, current_cpu, target_cpu, cpu_request, min_replica, Predicted_Demand):


    cpu_percentage = (current_cpu / cpu_request) * 100                                 # Calculating Percent CPU Utilization for the microservice deployment

    #******* Prediction of Resource Demand ***********
    

    Predicted_Demand = current_replicas
    Predicted_Time = Test_Time

    if (Test_Time > 10):                                                        # Prediction of Resource Demand starts after 10 seconds after starting load test

        with open('redis-cart_Demand_Model.pkl', 'rb') as f:                          # Loading of trained PROPHET model 
            redis_model = pickle.load(f)

        Predicted_Time = math.ceil (Test_Time/5)
        Predicted_Time = (Predicted_Time * 5) + 25                                  # Here you need to mention PREDICTION WINDOW SIZE (e.g., 25 seconds)               

        if Predicted_Time > 900:                                                    # 900 seconds is the total load test time (i.e., 900 seconds = 15 minutes)
            Predicted_Time = 900

        # Convert predicted time to datetime format
        Predicted_Time = pd.to_datetime(Predicted_Time, unit='s').strftime('1970-01-01 %H:%M:%S')                   #Date "1970-01-01" is the PROPHET's default settings (can be changed)
        
        # Create a future dataframe for the specified prediction time
        future = pd.DataFrame({'ds': [Predicted_Time]})

        # Add real-time regressor values to the future dataframe 
    
        futures  = future.copy() 
        futures.index = pd.to_datetime(futures.ds)   
        regressors = pd.concat([current_replicas, cpu_percentage, target_cpu], axis=1)    
        futures = futures.merge(regressors, left_index=True, right_index=True)     
        futures = futures.reset_index(drop = True)
        

        # Make prediction

        forecast = redis_model.predict(futures)

        Predicted_Demand = math.ceil(forecast ['yhat'])                                            # yhat is the desired replica count (i.e., Resource Demand)

        
    
    #******* Violation Detection and Corresponding Scaling Decision  ***********

    # Scale up

    if (Predicted_Demand > current_replicas) and (Predicted_Demand >= min_replica):
        scaling_action = "scale up"
        

    # Scale down


    elif (Predicted_Demand < current_replicas) and (Predicted_Demand >= min_replica):
        scaling_action = "scale down"
        

    #  No Scaling 

    else:
        scaling_action = "no scale"


    return cpu_percentage, scaling_action, Predicted_Demand, Predicted_Time



#***************************************************************************** Plan Component *****************************************************************************************

def Plan(microservice_name, Predicted_Time, Predicted_Demand, scaling_action, current_replicas, max_replica, cpu_request):

    
    with open(f'{microservice_name}.txt', 'w') as file:
        file.write(f'{Predicted_Demand}\n')
        file.write(f'{scaling_action}\n')
        file.write(f'{current_replicas}\n')
        file.write(f'{max_replica}\n')
        file.write(f'{cpu_request}\n')
        
    if (Predicted_Demand > max_replica):
        Application_Resource_Manager (Predicted_Time)

    return


#***************************************************************************** Execute Component *****************************************************************************************

def Execute (microservice_name, Predicted_Demand):
    execute_command = f"kubectl scale deployment {microservice_name} --replicas={Predicted_Demand}"
    os.system(execute_command)

    #Turning the status Application Resource Manager off for next iteration
    with open('ARM_redis-cart.txt', 'w') as file:
        for _ in range(4):
            file.write('0\n')
    return


#********************************************************************** Microservice_Manager_Main_Function ********************************************************************************



def redis(Test_Time):
    
    microservice_name, current_replicas, current_cpu, target_cpu, cpu_request, max_replica, min_replica = Monitor ()
     
    cpu_percentage, scaling_action, Predicted_Demand, Predicted_Time = Analyse (Test_Time, current_replicas, current_cpu, target_cpu, cpu_request, min_replica)

    Plan (microservice_name, Predicted_Time, Predicted_Demand, scaling_action, current_replicas, max_replica, cpu_request)

    # Checking for Application Resource Manager Status

    with open('ARM_redis-cart.txt', 'r') as file:
        lines = file.readlines()
        ARM_status = int(lines[0].strip())
        if (ARM_status == 1):                                          # ARM Status = 1 means Application Resource Manager got activated and potentially can change the scaling action, resource capacity and demand.
            scaling_action = lines[1].strip()
            Predicted_Demand = int(lines[2].strip())
            max_replica = int(lines[3].strip())

    
    Execute (microservice_name, Predicted_Demand)
    
    
    #*********************************************************** Storing Data for each iteration in the Knowledge Base ****************************************************************

    workbook = load_workbook('./Knowledge Base/redis-cart.xlsx')
    sheet = workbook.active

    sheet.cell(row=1, column=1, value="Test Time (sec)")
    sheet.cell(row=1, column=2, value="CPU Usage Percentage")
    sheet.cell(row=1, column=3, value="Current Replicas")
    sheet.cell(row=1, column=4, value="Predicted Desired Replicas")
    sheet.cell(row=1, column=5, value="Max Replicas")
    sheet.cell(row=1, column=6, value="Scaling Decision")
    sheet.cell(row=1, column=8, value="Predicted Time")


    
    Test_Time_row_length = len(sheet['A']) + 1 
    CPU_usage_row_length = len(sheet['B']) + 1
    current_replica_row_length = len(sheet['C']) + 1
    Predicted_Demand_row_length = len(sheet['D']) + 1
    Max_Replicas_row_length = len(sheet['E']) + 1
    Scaling_Decision_row_length = len(sheet['F']) + 1
    Predicted_Time_row_length = len(sheet['H']) + 1
 
 


    sheet[f'A{Test_Time_row_length}'] = Test_Time
    sheet[f'B{CPU_usage_row_length}'] = cpu_percentage
    sheet[f'C{current_replica_row_length}'] = current_replicas
    sheet[f'D{Predicted_Demand_row_length}'] = Predicted_Demand
    sheet[f'E{Max_Replicas_row_length}'] = max_replica
    sheet[f'F{Scaling_Decision_row_length}'] = scaling_action
    sheet[f'H{Predicted_Time_row_length}'] = Predicted_Time

    workbook.save('./Knowledge Base/redis-cart.xlsx')

    return







